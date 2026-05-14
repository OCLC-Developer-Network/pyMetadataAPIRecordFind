#!/usr/bin/env python3
# Copyright 2024
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
OCLC ISBN Matcher - WorldCat Metadata API Version

This script reads ISBNs from a spreadsheet (Excel, CSV, or TSV), searches the
WorldCat Metadata API for matching records, and adds the OCLC numbers to a new
column in the output Excel file.

Features:
- Accepts Excel (.xlsx/.xls), UTF-8 CSV, UTF-8 TSV, or MARC (.mrc/.marc) input
- Handles multiple ISBN columns (XML ISBN, HC ISBN, PB ISBN, ePub ISBN, ePDF ISBN)
- Maps format types to appropriate itemSubType parameters for API calls
- Searches WorldCat Metadata API with OAuth 2.0 authentication
- Adds rate limiting and error handling
- Provides detailed logging and progress tracking
- Creates backup of original file
- Uses environment variables for secure credential management
"""

import openpyxl
import csv
import requests
from requests_oauthlib import OAuth2Session
from oauthlib.oauth2 import BackendApplicationClient
import time
import logging
import shutil
import argparse
from datetime import datetime
from typing import Optional, Dict, Any, List, Tuple
import sys
from pathlib import Path
import os
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('oclc_matcher.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class OCLCISBNMatcher:
    """Class to handle WorldCat Metadata API searches and tabular input (Excel, CSV, TSV)."""
    
    def __init__(self, base_url: Optional[str] = None, 
                 api_key: Optional[str] = None,
                 api_secret: Optional[str] = None,
                 oauth_token_url: Optional[str] = None,
                 api_logging: Optional[bool] = None,
                 timeout: Optional[int] = None,
                 rate_limit_delay: Optional[float] = None):
        """
        Initialize the OCLC ISBN Matcher with WorldCat Metadata API.
        
        Args:
            base_url: Base URL for the WorldCat Metadata API (defaults to env var or production URL)
            api_key: OCLC API key (defaults to OCLC_API_KEY env var)
            api_secret: OCLC API secret (defaults to OCLC_API_SECRET env var)
            oauth_token_url: OAuth token URL (defaults to OCLC_OAUTH_TOKEN_URL env var or default)
            api_logging: Whether to enable detailed API request/response logging (defaults to env var)
            timeout: Request timeout in seconds (defaults to API_TIMEOUT env var or 30)
            rate_limit_delay: Delay between requests in seconds (defaults to API_RATE_LIMIT_DELAY env var or 0.5)
        """
        # Load configuration from environment variables
        self.base_url = base_url or os.getenv('OCLC_API_BASE_URL', 'https://metadata.api.oclc.org')
        self.api_key = api_key or os.getenv('OCLC_API_KEY')
        self.api_secret = api_secret or os.getenv('OCLC_API_SECRET')
        self.oauth_token_url = oauth_token_url or os.getenv('OCLC_OAUTH_TOKEN_URL', 'https://oauth.oclc.org/token')
        self.api_logging = api_logging if api_logging is not None else os.getenv('API_LOGGING', 'true').lower() == 'true'
        self.timeout = timeout or int(os.getenv('API_TIMEOUT', '30'))
        self.rate_limit_delay = rate_limit_delay or float(os.getenv('API_RATE_LIMIT_DELAY', '0.5'))
        
        # Validate required credentials
        if not self.api_key or not self.api_secret:
            raise ValueError(
                "OCLC API credentials are required. "
                "Set OCLC_API_KEY and OCLC_API_SECRET environment variables "
                "or provide them as arguments. See .env.example for details."
            )
        
        # Initialize OAuth 2.0 client credentials flow
        self.client = BackendApplicationClient(client_id=self.api_key)
        self.oauth = OAuth2Session(client=self.client)
        
        # Get access token
        self._refresh_access_token()
        
        # Statistics tracking
        self.stats = {
            'total_processed': 0,
            'successful_matches': 0,
            'api_errors': 0,
            'empty_isbns': 0,
            'no_matches': 0,
            'lcsh_found': 0,
            'lcsh_not_found': 0,
            'api_requests': 0,
            'api_responses': 0
        }
    
    def _refresh_access_token(self):
        """
        Refresh the OAuth 2.0 access token using client credentials flow.
        """
        try:
            # OCLC OAuth requires Basic Auth with client_id:client_secret
            # and grant_type=client_credentials in the body
            import base64
            auth_string = f"{self.api_key}:{self.api_secret}"
            auth_bytes = auth_string.encode('ascii')
            auth_b64 = base64.b64encode(auth_bytes).decode('ascii')
            
            headers = {
                'Authorization': f'Basic {auth_b64}',
                'Content-Type': 'application/x-www-form-urlencoded',
                'Accept': 'application/json'
            }
            
            data = {
                'grant_type': 'client_credentials',
                'scope': 'WorldCatMetadataAPI'
            }
            
            if self.api_logging:
                logger.info(f"OAuth Token Request - URL: {self.oauth_token_url}")
                logger.info(f"OAuth Token Request - Headers: {dict(headers)}")
                logger.info(f"OAuth Token Request - Data: {data}")
            
            response = requests.post(
                self.oauth_token_url,
                headers=headers,
                data=data,
                timeout=self.timeout
            )
            
            if self.api_logging:
                logger.info(f"OAuth Token Response - Status: {response.status_code}")
                logger.info(f"OAuth Token Response - Headers: {dict(response.headers)}")
                logger.info(f"OAuth Token Response - Body: {response.text[:500]}")
            
            response.raise_for_status()
            
            token_data = response.json()
            
            # Check if access_token is in the response
            if 'access_token' not in token_data:
                logger.error(f"OAuth response missing access_token. Full response: {token_data}")
                raise ValueError(
                    f"OAuth response missing access_token. Response keys: {list(token_data.keys())}"
                )
            
            self.access_token = token_data['access_token']
            
            if not self.access_token:
                logger.error(f"OAuth response has empty access_token. Full response: {token_data}")
                raise ValueError("OAuth response has empty access_token")
            
            logger.info("Successfully obtained OAuth access token")
            
            # Log token expiration if available
            if 'expires_in' in token_data:
                logger.debug(f"Token expires in {token_data['expires_in']} seconds")
                
        except requests.exceptions.RequestException as e:
            logger.error(f"HTTP error during OAuth token request: {e}")
            if hasattr(e, 'response') and e.response is not None:
                logger.error(f"Response status: {e.response.status_code}")
                logger.error(f"Response headers: {dict(e.response.headers)}")
                try:
                    error_body = e.response.text
                    logger.error(f"Response body: {error_body[:500]}")
                except:
                    logger.error("Could not read response body")
            raise ValueError(f"Authentication failed: {e}")
        except Exception as e:
            logger.error(f"Failed to obtain OAuth access token: {e}")
            logger.error(f"Exception type: {type(e).__name__}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            raise ValueError(f"Authentication failed: {e}")
    
    def _get_headers(self) -> Dict[str, str]:
        """
        Get headers for API requests including OAuth token.
        
        Returns:
            Dictionary of HTTP headers
        """
        return {
            'Accept': 'application/json',
            'Authorization': f'Bearer {self.access_token}'
        }
    
    def print_api_statistics(self):
        """Print API usage statistics."""
        logger.info("=" * 60)
        logger.info("API USAGE STATISTICS")
        logger.info("=" * 60)
        logger.info(f"Total API Requests: {self.stats['api_requests']}")
        logger.info(f"Total API Responses: {self.stats['api_responses']}")
        logger.info(f"API Errors: {self.stats['api_errors']}")
        if self.stats['api_requests'] > 0:
            success_rate = ((self.stats['api_requests'] - self.stats['api_errors']) / self.stats['api_requests']) * 100
            logger.info(f"API Success Rate: {success_rate:.1f}%")
        logger.info("=" * 60)
    
    def search_by_isbns(self, isbns: list, format_type: str = None) -> dict:
        """
        Search OCLC API for records by multiple ISBNs using OR query.
        
        Args:
            isbns: List of ISBNs to search for
            format_type: Format type to map to itemSubType parameter
            
        Returns:
            Dictionary mapping ISBN to OCLC number (if found)
        """
        try:
            # Clean and validate ISBNs
            clean_isbns = []
            isbn_mapping = {}  # Maps clean ISBN back to original
            
            for isbn in isbns:
                if not isbn or str(isbn).strip() == '':
                    continue
                    
                clean_isbn = str(isbn).replace('-', '').replace(' ', '').strip()
                
                # Validate ISBN length (should be 10 or 13 digits)
                if not clean_isbn.isdigit() or len(clean_isbn) not in [10, 13]:
                    logger.warning(f"Invalid ISBN format: {isbn}")
                    continue
                
                clean_isbns.append(clean_isbn)
                isbn_mapping[clean_isbn] = isbn
            
            if not clean_isbns:
                logger.warning("No valid ISBNs provided")
                return {}
            
            # Construct OR query for multiple ISBNs
            query_parts = [f"bn:{isbn}" for isbn in clean_isbns]
            query = " OR ".join(query_parts)
            
            # API endpoint for WorldCat Metadata API search (using brief-bibs endpoint)
            url = f"{self.base_url}/worldcat/search/brief-bibs"
            
            # Determine whether to use itemType or itemSubType based on format
            if format_type is None:
                # No format specified, don't send itemType or itemSubType
                params = {
                    'q': query,
                    'groupRelatedEditions': 'true',
                    'inCatalogLanguage': 'eng',
                    'orderBy': 'mostWidelyHeld',
                    'limit': 1  # Only need one result since all ISBNs are for the same work
                }
            elif self._should_use_item_type(format_type):
                # Use itemType parameter for formats that don't support itemSubType
                item_type = self._get_item_type_for_format(format_type)
                params = {
                    'q': query,
                    'groupRelatedEditions': 'true',
                    'inCatalogLanguage': 'eng',
                    'orderBy': 'mostWidelyHeld',
                    'itemType': item_type,
                    'limit': 1  # Only need one result since all ISBNs are for the same work
                }
            else:
                # Use itemSubType parameter for supported formats
                item_sub_type = self._map_format_to_item_sub_type(format_type)
                params = {
                    'q': query,
                    'groupRelatedEditions': 'true',
                    'inCatalogLanguage': 'eng',
                    'orderBy': 'mostWidelyHeld',
                    'itemSubType': item_sub_type,
                    'limit': 1  # Only need one result since all ISBNs are for the same work
                }
            
            # Get headers with OAuth token
            headers = self._get_headers()
            
            # Log API request details
            if self.api_logging:
                logger.info(f"API Request - ISBN Search")
                logger.info(f"  URL: {url}")
                logger.info(f"  Query: {query}")
                logger.info(f"  Parameters: {params}")
                logger.info(f"  Headers: {headers}")
            
            self.stats['api_requests'] += 1
            response = requests.get(url, params=params, headers=headers, timeout=self.timeout)
            
            # Handle 401 Unauthorized - token may have expired
            if response.status_code == 401:
                logger.warning("Received 401 Unauthorized, refreshing access token...")
                self._refresh_access_token()
                headers = self._get_headers()
                response = requests.get(url, params=params, headers=headers, timeout=self.timeout)
            
            # Log response details
            if self.api_logging:
                logger.info(f"API Response - ISBN Search")
                logger.info(f"  Status Code: {response.status_code}")
                logger.info(f"  Response Headers: {dict(response.headers)}")
                logger.info(f"  Response Size: {len(response.content)} bytes")
            
            self.stats['api_responses'] += 1
            
            response.raise_for_status()
            
            data = response.json()
            
            # Log response content (truncated for large responses)
            if self.api_logging:
                response_content = str(data)
                if len(response_content) > 1000:
                    response_content = response_content[:1000] + "... [truncated]"
                logger.info(f"  Response Content: {response_content}")
            
            # Process results and map back to original ISBNs
            # brief-bibs endpoint returns 'briefRecords' array, not 'bibRecords'
            results = {}
            brief_records = data.get('briefRecords', [])
            
            if brief_records:
                brief_record = brief_records[0]
                # Extract OCLC number directly from brief record (not nested in identifier)
                oclc_number = brief_record.get('oclcNumber')
                
                if oclc_number:
                    # Check for LCSH subjects by fetching full record from /bibs endpoint
                    has_lcsh = self._check_lcsh_in_bib_record(oclc_number)
                    
                    # Since all ISBNs in the query are for the same work,
                    # we can associate the found OCLC number and LCSH status with all of them
                    for original_isbn in isbns:
                        results[original_isbn] = {
                            'oclc_number': oclc_number,
                            'has_lcsh': has_lcsh
                        }
                    
                    logger.debug(f"Found OCLC number: {oclc_number} with LCSH: {has_lcsh}")
            
            logger.debug(f"Found {len(results)} matches out of {len(clean_isbns)} ISBNs")
            return results
            
        except requests.exceptions.RequestException as e:
            logger.error(f"API request failed for ISBNs {isbns}: {e}")
            if hasattr(e, 'response') and e.response is not None:
                logger.error(f"  Response Status: {e.response.status_code}")
                logger.error(f"  Response Headers: {dict(e.response.headers)}")
                try:
                    error_content = e.response.text
                    if len(error_content) > 500:
                        error_content = error_content[:500] + "... [truncated]"
                    logger.error(f"  Response Content: {error_content}")
                except:
                    logger.error(f"  Could not read response content")
            self.stats['api_errors'] += 1
            return {}
        except Exception as e:
            logger.error(f"Unexpected error searching for ISBNs {isbns}: {e}")
            return {}

    def search_by_title_author_publisher(self, title: str, author: str, publisher: str, 
                                       publication_date: str, format_type: str = None, other_identifier: str = None) -> dict:
        """
        Search OCLC Discovery API using title, author, publisher, and other identifier when no ISBN is available.
        First tries with publication date, then retries without date if no results found.
        
        Args:
            title: Book title
            author: Author name
            publisher: Publisher name
            publication_date: Publication date (YYYY format)
            format_type: Format type to map to itemSubType parameter
            other_identifier: Other identifier (e.g., from MARC 024$a)
            
        Returns:
            Dictionary with search results
        """
        # Build search query components
        query_parts = []
        
        if title and str(title).strip():
            # Escape special characters for title search
            clean_title = str(title).strip().replace('"', '\\"')
            query_parts.append(f'te:{clean_title}')
        
        if author and str(author).strip():
            # Escape special characters and wrap in quotes for exact phrase matching
            clean_author = str(author).strip().replace('"', '\\"')
            query_parts.append(f'au:"{clean_author}"')
        
        if publisher and str(publisher).strip():
            # Escape special characters and wrap in quotes for exact phrase matching
            clean_publisher = str(publisher).strip().replace('"', '\\"')
            query_parts.append(f'pb:"{clean_publisher}"')
        
        if other_identifier and str(other_identifier).strip():
            # Escape special characters and wrap in quotes for exact phrase matching
            clean_other_id = str(other_identifier).strip().replace('"', '\\"')
            query_parts.append(f'sn:"{clean_other_id}"')
        
        if not query_parts:
            logger.warning("No searchable fields provided (title, author, publisher, other identifier)")
            return {'oclc_number': None, 'has_lcsh': False}
        
        # Join with AND operators
        query = " AND ".join(query_parts)
        logger.debug(f"Searching by title/author/publisher with query: {query}")
        
        # Try search with publication date first, then without if no results
        search_attempts = []
        
        # First attempt: with publication date (if available)
        if publication_date and str(publication_date).strip():
            import re
            year_match = re.search(r'\b(19|20)\d{2}\b', str(publication_date))
            if year_match:
                search_attempts.append(('with date', year_match.group()))
        
        # Second attempt: without publication date
        search_attempts.append(('without date', None))
        
        for attempt_name, date_value in search_attempts:
            try:
                # Determine whether to use itemType or itemSubType based on format
                if format_type is None:
                    # No format specified, don't send itemType or itemSubType
                    params = {
                        'q': query,
                        'groupRelatedEditions': 'true',
                        'inCatalogLanguage': 'eng',
                        'orderBy': 'mostWidelyHeld',
                        'limit': 1
                    }
                elif self._should_use_item_type(format_type):
                    # Use itemType parameter for formats that don't support itemSubType
                    item_type = self._get_item_type_for_format(format_type)
                    params = {
                        'q': query,
                        'groupRelatedEditions': 'true',
                        'inCatalogLanguage': 'eng',
                        'orderBy': 'mostWidelyHeld',
                        'itemType': item_type,
                        'limit': 1
                    }
                else:
                    # Use itemSubType parameter for supported formats
                    item_sub_type = self._map_format_to_item_sub_type(format_type)
                    params = {
                        'q': query,
                        'groupRelatedEditions': 'true',
                        'inCatalogLanguage': 'eng',
                        'orderBy': 'mostWidelyHeld',
                        'itemSubType': item_sub_type,
                        'limit': 1
                    }
                
                # Add publication date if available for this attempt
                if date_value:
                    params['datePublished'] = date_value
                    logger.debug(f"Added datePublished parameter: {params['datePublished']}")
                
                # Get headers with OAuth token
                headers = self._get_headers()
                
                # Log API request details
                # API endpoint for WorldCat Metadata API search (using brief-bibs endpoint)
                url = f"{self.base_url}/worldcat/search/brief-bibs"
                if self.api_logging:
                    logger.info(f"API Request - Alternative Search ({attempt_name})")
                    logger.info(f"  URL: {url}")
                    logger.info(f"  Query: {query}")
                    logger.info(f"  Parameters: {params}")
                    logger.info(f"  Headers: {headers}")
                
                self.stats['api_requests'] += 1
                response = requests.get(url, params=params, headers=headers, timeout=self.timeout)
                
                # Handle 401 Unauthorized - token may have expired
                if response.status_code == 401:
                    logger.warning("Received 401 Unauthorized, refreshing access token...")
                    self._refresh_access_token()
                    headers = self._get_headers()
                    response = requests.get(url, params=params, headers=headers, timeout=self.timeout)
                
                # Log response details
                if self.api_logging:
                    logger.info(f"API Response - Alternative Search ({attempt_name})")
                    logger.info(f"  Status Code: {response.status_code}")
                    logger.info(f"  Response Headers: {dict(response.headers)}")
                    logger.info(f"  Response Size: {len(response.content)} bytes")
                
                self.stats['api_responses'] += 1
                
                response.raise_for_status()
                
                data = response.json()
                
                # Log response content (truncated for large responses)
                if self.api_logging:
                    response_content = str(data)
                    if len(response_content) > 1000:
                        response_content = response_content[:1000] + "... [truncated]"
                    logger.info(f"  Response Content: {response_content}")
                
                # Process results
                # brief-bibs endpoint returns 'briefRecords' array, not 'bibRecords'
                brief_records = data.get('briefRecords', [])
                if brief_records:
                    brief_record = brief_records[0]  # Get first result
                    # Extract OCLC number directly from brief record (not nested in identifier)
                    oclc_number = brief_record.get('oclcNumber')
                    # Check for LCSH subjects by fetching full record from /bibs endpoint
                    has_lcsh = self._check_lcsh_in_bib_record(oclc_number) if oclc_number else False
                    
                    logger.debug(f"Found match {attempt_name}: OCLC {oclc_number}, LCSH: {has_lcsh}")
                    return {
                        'oclc_number': oclc_number,
                        'has_lcsh': has_lcsh
                    }
                else:
                    logger.debug(f"No results found for title/author/publisher search {attempt_name}")
                    # Continue to next attempt if this one had no results
                    continue
                    
            except requests.exceptions.RequestException as e:
                logger.error(f"API request failed for title/author/publisher search ({attempt_name}): {e}")
                if hasattr(e, 'response') and e.response is not None:
                    logger.error(f"  Response Status: {e.response.status_code}")
                    logger.error(f"  Response Headers: {dict(e.response.headers)}")
                    try:
                        error_content = e.response.text
                        if len(error_content) > 500:
                            error_content = error_content[:500] + "... [truncated]"
                        logger.error(f"  Response Content: {error_content}")
                    except:
                        logger.error(f"  Could not read response content")
                self.stats['api_errors'] += 1
                # Continue to next attempt on API error
                continue
            except Exception as e:
                logger.error(f"Unexpected error in title/author/publisher search ({attempt_name}): {e}")
                # Continue to next attempt on unexpected error
                continue
        
        # If we get here, no attempts succeeded
        logger.debug("No results found for title/author/publisher search after all attempts")
        return {
            'oclc_number': None,
            'has_lcsh': False
        }
    
    def _check_lcsh_in_bib_record(self, oclc_number: str) -> bool:
        """
        Check if a bib record contains Library of Congress Subject Headings (LCSH).
        Fetches the full bibliographic record from the /bibs endpoint to check for LCSH.
        
        Args:
            oclc_number: OCLC number of the record to check
            
        Returns:
            True if the record contains LCSH subjects, False otherwise
        """
        if not oclc_number:
            logger.debug("No OCLC number provided for LCSH check")
            return False
            
        try:
            # Fetch full bibliographic record from /bibs endpoint
            url = f"{self.base_url}/worldcat/bibs/{oclc_number}"
            headers = self._get_headers()
            
            if self.api_logging:
                logger.info(f"API Request - Fetch Full Bib for LCSH Check")
                logger.info(f"  URL: {url}")
                logger.info(f"  Headers: {headers}")
            
            self.stats['api_requests'] += 1
            response = requests.get(url, headers=headers, timeout=self.timeout)
            
            # Handle 401 Unauthorized - token may have expired
            if response.status_code == 401:
                logger.warning("Received 401 Unauthorized while fetching full bib, refreshing access token...")
                self._refresh_access_token()
                headers = self._get_headers()
                response = requests.get(url, headers=headers, timeout=self.timeout)
            
            if self.api_logging:
                logger.info(f"API Response - Fetch Full Bib for LCSH Check")
                logger.info(f"  Status Code: {response.status_code}")
                logger.info(f"  Response Headers: {dict(response.headers)}")
                logger.info(f"  Response Size: {len(response.content)} bytes")
            
            self.stats['api_responses'] += 1
            
            # If record not found or other error, return False
            if response.status_code == 404:
                logger.debug(f"Record {oclc_number} not found for LCSH check")
                return False
            
            response.raise_for_status()
            
            bib_record = response.json()
            
            # Log response content (truncated for large responses)
            if self.api_logging:
                response_content = str(bib_record)
                if len(response_content) > 1000:
                    response_content = response_content[:1000] + "... [truncated]"
                logger.info(f"  Response Content: {response_content}")
            
            # Check for LCSH subjects in the full record
            subjects = bib_record.get('subjects', [])
            
            # Check if any subject has LCSH vocabulary
            for subject in subjects:
                vocabulary = subject.get('vocabulary', '')
                if vocabulary and 'LIBRARY OF CONGRESS SUBJECT HEADINGS' in vocabulary.upper():
                    logger.debug(f"Found LCSH subject: {vocabulary}")
                    return True
            
            logger.debug(f"No LCSH subjects found in record {oclc_number}")
            return False
            
        except requests.exceptions.RequestException as e:
            logger.error(f"API request failed while checking LCSH for OCLC {oclc_number}: {e}")
            if hasattr(e, 'response') and e.response is not None:
                logger.error(f"  Response Status: {e.response.status_code}")
                logger.error(f"  Response Headers: {dict(e.response.headers)}")
                try:
                    error_content = e.response.text
                    if len(error_content) > 500:
                        error_content = error_content[:500] + "... [truncated]"
                    logger.error(f"  Response Content: {error_content}")
                except:
                    logger.error(f"  Could not read response content")
            # Return False on error rather than raising
            return False
        except Exception as e:
            logger.error(f"Unexpected error checking LCSH in bib record {oclc_number}: {e}")
            return False
    
    def search_by_isbn(self, isbn: str) -> Optional[str]:
        """
        Search OCLC API for a single ISBN (backward compatibility).
        
        Args:
            isbn: ISBN to search for
            
        Returns:
            OCLC number if found, None otherwise
        """
        results = self.search_by_isbns([isbn])
        result = results.get(isbn)
        if result and isinstance(result, dict):
            return result.get('oclc_number')
        return result
    
    def _map_format_to_item_sub_type(self, format_type: str) -> str:
        """
        Map format type to OCLC itemSubType parameter.
        
        Args:
            format_type: Format type from Excel file
            
        Returns:
            Corresponding itemSubType parameter for OCLC API
        """
        if not format_type:
            return 'book-digital'  # Default fallback
        
        format_mapping = {
            'book-print': 'book-printbook',
            'book-digital': 'book-digital', 
            'book-largeprint': 'book-largeprint',
            'print': 'book-printbook',
            'hardcover': 'book-printbook',
            'paperback': 'book-printbook',
            'video': 'video',
            'audiobook': 'audiobook',
            'music': 'music'
        }
        
        # Normalize format type (remove extra spaces, convert to lowercase)
        normalized_format = str(format_type).strip().lower()
        
        return format_mapping.get(normalized_format, 'book-digital')
    
    def _should_use_item_type(self, format_type: str) -> bool:
        """
        Determine if the format should be sent as itemType parameter instead of itemSubType.
        
        Based on API testing, only certain formats work with itemSubType:
        - book-digital and book-largeprint work with itemSubType
        - Other formats should use itemType parameter
        
        Args:
            format_type: Format type from Excel file
            
        Returns:
            True if should use itemType, False if should use itemSubType
        """
        if not format_type:
            return False  # Default to itemSubType
        
        # Normalize format type
        normalized_format = str(format_type).strip().lower()
        
        # Only book-digital, book-largeprint, and book-print work with itemSubType
        item_sub_type_formats = {
            'book-digital',
            'book-largeprint',
            'book-large-print',
            'large-print',
            'largeprint',
            'book-print',
            'print',
            'hardcover',
            'paperback',
            'ebook',
            'e-book',
            'electronic',
            'digital'
        }
        
        # Check if this format should use itemSubType
        if normalized_format in item_sub_type_formats:
            return False  # Use itemSubType
        
        # Check for partial matches
        for format_key in item_sub_type_formats:
            if format_key in normalized_format or normalized_format in format_key:
                return False  # Use itemSubType
        
        # All other formats should use itemType
        return True
    
    def _get_item_type_for_format(self, format_type: str) -> str:
        """
        Get the appropriate itemType for formats that don't support itemSubType.
        
        Args:
            format_type: Format type from Excel file
            
        Returns:
            Corresponding itemType parameter for OCLC API
        """
        if not format_type:
            return 'book'  # Default fallback
        
        # Normalize format type
        normalized_format = str(format_type).strip().lower()
        
        # Map formats to their appropriate itemType
        item_type_mapping = {
            'video': 'video',
            'video-recording': 'video',
            'motion-picture': 'video',
            'film': 'video',
            'dvd': 'video',
            'blu-ray': 'video',
            'audiobook': 'audiobook',
            'audio-book': 'audiobook',
            'audio': 'audiobook',
            'sound-recording': 'audiobook',
            'spoken-word': 'audiobook',
            'music': 'music',
            'musical-recording': 'music',
            'sound-recording-music': 'music',
            'cd': 'music',
            'vinyl': 'music',
            'record': 'music',
            'compfile': 'compfile',
            'computer-file': 'compfile',
            'computer': 'compfile',
            'game': 'game',
            'computer-game': 'game',
            'video-game': 'game',
            'print': 'book',
            'hardcover': 'book',
            'paperback': 'book'
        }
        
        # Direct lookup first
        if normalized_format in item_type_mapping:
            return item_type_mapping[normalized_format]
        
        # Partial matching for variations
        for key, item_type in item_type_mapping.items():
            if key in normalized_format or normalized_format in key:
                return item_type
        
        # Fallback based on common patterns
        if any(term in normalized_format for term in ['video', 'film', 'movie', 'dvd', 'blu-ray']):
            return 'video'
        elif any(term in normalized_format for term in ['audio', 'sound', 'spoken', 'audiobook']):
            return 'audiobook'
        elif any(term in normalized_format for term in ['music', 'song', 'album', 'cd', 'vinyl']):
            return 'music'
        elif any(term in normalized_format for term in ['game', 'computer-game', 'video-game']):
            return 'game'
        elif any(term in normalized_format for term in ['computer', 'compfile', 'software', 'program']):
            return 'compfile'
        elif any(term in normalized_format for term in ['book', 'print', 'hardcover', 'paperback']):
            return 'book'
        
        # Final fallback
        return 'book'
    
    def find_format_column(self, worksheet) -> Optional[int]:
        """
        Find the format column in the worksheet.
        
        Args:
            worksheet: OpenPyXL worksheet object
            
        Returns:
            Column index of format column, or None if not found
        """
        # Look for common format column names (case-insensitive matching)
        format_column_names = [
            'Format', 'Format Type', 'Type'
        ]
        # Normalize to lowercase for case-insensitive comparison
        format_column_names_lower = {name.lower() for name in format_column_names}
        
        # Check the first row for column headers
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value:
                cell_value_str = str(cell_value).strip()
                if cell_value_str.lower() in format_column_names_lower:
                    logger.info(f"Found format column: {cell_value} at column {col}")
                    return col
        
        logger.warning("No format column found - will use default itemSubType")
        return None
    
    def find_description_column(self, worksheet) -> Optional[int]:
        """
        Find the description column in the worksheet.
        
        Args:
            worksheet: OpenPyXL worksheet object
            
        Returns:
            Column index of description column, or None if not found
        """
        # Look for common description column names (case-insensitive matching)
        description_column_names = [
            'Description', 'Physical Description', 'PhysicalDesc', 'Desc'
        ]
        # Normalize to lowercase for case-insensitive comparison
        description_column_names_lower = {name.lower() for name in description_column_names}
        
        # Check the first row for column headers
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value:
                cell_value_str = str(cell_value).strip()
                if cell_value_str.lower() in description_column_names_lower:
                    logger.info(f"Found description column: {cell_value} at column {col}")
                    return col
        
        logger.debug("No description column found")
        return None
    
    def determine_final_format(self, format_value: str, description_value: str = None) -> str:
        """
        Determine the final format to use, checking description field last.
        If description contains 'computer' and 'game', set format to 'game'.
        If description contains 'computer' (but not 'game'), set format to 'compfile'.
        If description contains 'audio media player', set format to None.
        
        Args:
            format_value: Format value from format column
            description_value: Description value from description column
            
        Returns:
            Final format to use for API calls, or None if audio media player
        """
        # First, use the format value if available
        if format_value and str(format_value).strip():
            format_str = str(format_value).strip()
        else:
            format_str = None
        
        # Check description field last for specific overrides
        if description_value and str(description_value).strip():
            description_str = str(description_value).strip().lower()
            
            # Check for computer + game combination first (more specific)
            if 'computer' in description_str and 'game' in description_str:
                logger.info(f"Description contains 'computer' and 'game', setting format to 'game' (was: {format_str})")
                return 'game'
            elif 'computer' in description_str:
                logger.info(f"Description contains 'computer', setting format to 'compfile' (was: {format_str})")
                return 'compfile'
            elif 'audio media player' in description_str:
                logger.info(f"Description contains 'audio media player', setting format to None (was: {format_str})")
                return None
        
        # Return the original format value or default
        return format_str if format_str else 'book-digital'

    def find_column_by_name(self, worksheet, column_names: list) -> Optional[int]:
        """
        Find a column by name in the worksheet (case-insensitive matching).
        
        Args:
            worksheet: OpenPyXL worksheet object
            column_names: List of possible column names to search for
            
        Returns:
            Column index of the column, or None if not found
        """
        # Normalize to lowercase for case-insensitive comparison
        column_names_lower = {name.lower() if isinstance(name, str) else str(name).lower() for name in column_names}
        
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value:
                cell_value_str = str(cell_value).strip()
                if cell_value_str.lower() in column_names_lower:
                    logger.debug(f"Found column: {cell_value} at column {col}")
                    return col
        
        return None
    
    def find_isbn_columns(self, worksheet) -> List[Tuple[int, str]]:
        """
        Find all columns containing ISBNs by checking if column name contains "ISBN" 
        (case-insensitive substring matching).
        
        Args:
            worksheet: OpenPyXL worksheet object
            
        Returns:
            List of tuples (column_index, column_name) containing ISBNs
        """
        isbn_columns = []
        
        # Check the first row for column headers
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value:
                cell_value_str = str(cell_value).strip()
                # Check if column name contains "ISBN" (case-insensitive)
                if 'isbn' in cell_value_str.lower():
                    isbn_columns.append((col, cell_value_str))
                    logger.info(f"Found ISBN column: {cell_value} at column {col}")
        
        if not isbn_columns:
            logger.warning("No ISBN column headers found")
        
        return isbn_columns
    
    def create_backup(self, input_file: str) -> str:
        """Create a backup of the input file."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = f"{input_file}.backup_{timestamp}"
        shutil.copy2(input_file, backup_file)
        logger.info(f"Created backup: {backup_file}")
        return backup_file

    def _delimited_file_to_workbook(self, input_file: str, delimiter: str) -> openpyxl.Workbook:
        """
        Load a UTF-8 (with optional BOM) delimited file into a new single-sheet workbook.

        Args:
            input_file: Path to the CSV or TSV file
            delimiter: Field delimiter (comma or tab)

        Returns:
            New workbook with sheet data matching the delimited file rows.
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        with open(input_file, newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, value in enumerate(row, start=1):
                    worksheet.cell(
                        row=row_idx,
                        column=col_idx,
                        value=None if value == "" else value,
                    )
        return workbook

    def _process_workbook(self, workbook: openpyxl.Workbook, output_file: str) -> None:
        """
        Run OCLC matching on the active sheet and save to an Excel file.

        Args:
            workbook: Loaded or constructed workbook (active sheet is processed)
            output_file: Path to the output .xlsx file
        """
        worksheet = workbook.active

        # Find all ISBN columns
        isbn_columns = self.find_isbn_columns(worksheet)
        
        if not isbn_columns:
            raise ValueError("No ISBN columns found in the workbook")
        
        # Find format column
        format_col = self.find_format_column(worksheet)
        
        # Find description column
        description_col = self.find_description_column(worksheet)
        
        # Find additional columns for alternative search
        title_col = self.find_column_by_name(worksheet, ['Title', 'title', 'TITLE'])
        author_col = self.find_column_by_name(worksheet, ['Author', 'author', 'AUTHOR'])
        publisher_col = self.find_column_by_name(worksheet, ['Publisher', 'publisher', 'PUBLISHER'])
        pub_date_col = self.find_column_by_name(worksheet, ['Publication Date', 'publication_date', 'PublicationDate', 'Date', 'date'])
        other_id_col_input = self.find_column_by_name(worksheet, ['Other Identifier', 'other_identifier', 'OtherIdentifier', 'Other ID', 'other_id'])
        
        # Find the last column to add our new columns
        last_col = worksheet.max_column
        oclc_col = last_col + 1
        lcsh_col = last_col + 2
        other_id_col = last_col + 3
        
        # Add headers for new columns
        worksheet.cell(row=1, column=oclc_col, value='matchingOCLCNumber')
        worksheet.cell(row=1, column=lcsh_col, value='hasLCSHSubjects')
        worksheet.cell(row=1, column=other_id_col, value='Other Identifier')
        
        # Process each row individually
        total_rows = worksheet.max_row
        logger.info(f"Processing {total_rows - 1} records...")
        
        start_time = time.time()
        
        for row in range(2, total_rows + 1):  # Start from row 2 (skip header)
            row_isbns = []
            
            # Collect all ISBNs from this row
            for col_idx, col_name in isbn_columns:
                isbn_cell = worksheet.cell(row=row, column=col_idx)
                isbn = isbn_cell.value
                
                # Skip if ISBN is empty or None
                if not isbn or str(isbn).strip() == '':
                    continue
                
                isbn_str = str(isbn).strip()
                row_isbns.append(isbn_str)
            
            if not row_isbns:
                # No ISBNs in this row - try alternative search using title/author/publisher
                logger.info(f"Row {row}: No ISBNs found, attempting alternative search")
                
                # Get format and description values for this row
                format_value = None
                if format_col:
                    format_cell = worksheet.cell(row=row, column=format_col)
                    format_value = format_cell.value
                    if format_value:
                        format_value = str(format_value).strip()
                
                description_value = None
                if description_col:
                    description_cell = worksheet.cell(row=row, column=description_col)
                    description_value = description_cell.value
                    if description_value:
                        description_value = str(description_value).strip()
                
                # Determine final format (checking description last)
                final_format = self.determine_final_format(format_value, description_value)
                
                # Get title, author, publisher, and publication date
                title = None
                if title_col:
                    title_cell = worksheet.cell(row=row, column=title_col)
                    title = title_cell.value
                    if title:
                        title = str(title).strip()
                
                author = None
                if author_col:
                    author_cell = worksheet.cell(row=row, column=author_col)
                    author = author_cell.value
                    if author:
                        author = str(author).strip()
                
                publisher = None
                if publisher_col:
                    publisher_cell = worksheet.cell(row=row, column=publisher_col)
                    publisher = publisher_cell.value
                    if publisher:
                        publisher = str(publisher).strip()
                
                pub_date = None
                if pub_date_col:
                    pub_date_cell = worksheet.cell(row=row, column=pub_date_col)
                    pub_date = pub_date_cell.value
                    if pub_date:
                        pub_date = str(pub_date).strip()
                
                # Extract other identifier
                other_identifier = None
                if other_id_col_input:
                    other_id_cell = worksheet.cell(row=row, column=other_id_col_input)
                    other_identifier = other_id_cell.value
                    if other_identifier:
                        other_identifier = str(other_identifier).strip()
                
                # Search using title/author/publisher/other identifier
                format_display = final_format if final_format is not None else 'None (no itemType/itemSubType)'
                other_id_display = f", Other ID: '{other_identifier}'" if other_identifier else ""
                logger.info(f"Row {row}: Searching by title/author/publisher with format '{format_display}' - Title: '{title}', Author: '{author}', Publisher: '{publisher}', Date: '{pub_date}'{other_id_display}")
                result = self.search_by_title_author_publisher(title, author, publisher, pub_date, final_format, other_identifier)
                
                oclc_number = result.get('oclc_number')
                has_lcsh = result.get('has_lcsh', False)
                
                # Add OCLC number to the new column
                oclc_cell = worksheet.cell(row=row, column=oclc_col)
                oclc_cell.value = oclc_number
                
                # Add LCSH result to the new column
                lcsh_cell = worksheet.cell(row=row, column=lcsh_col)
                lcsh_cell.value = has_lcsh
                
                # Add other identifier to the new column (copy from input if available)
                other_id_cell = worksheet.cell(row=row, column=other_id_col)
                if other_id_col_input:
                    other_id_value = worksheet.cell(row=row, column=other_id_col_input).value
                    other_id_cell.value = other_id_value
                else:
                    other_id_cell.value = ''
                
                # Update LCSH statistics if we found an OCLC number
                if oclc_number:
                    if has_lcsh:
                        self.stats['lcsh_found'] += 1
                    else:
                        self.stats['lcsh_not_found'] += 1
                
                # Update statistics
                self.stats['total_processed'] += 1
                if oclc_number:
                    self.stats['successful_matches'] += 1
                    logger.info(f"Row {row}: Found match via alternative search -> OCLC: {oclc_number}, LCSH: {has_lcsh}")
                else:
                    self.stats['no_matches'] += 1
                    logger.warning(f"Row {row}: No match found via alternative search")
                
                # Add a small delay to be respectful to the API
                time.sleep(self.rate_limit_delay)
                continue
            
            # Get format and description values for this row
            format_value = None
            if format_col:
                format_cell = worksheet.cell(row=row, column=format_col)
                format_value = format_cell.value
                if format_value:
                    format_value = str(format_value).strip()
            
            description_value = None
            if description_col:
                description_cell = worksheet.cell(row=row, column=description_col)
                description_value = description_cell.value
                if description_value:
                    description_value = str(description_value).strip()
            
            # Determine final format (checking description last)
            final_format = self.determine_final_format(format_value, description_value)
            
            # Search for OCLC number using OR query for all ISBNs in this row
            format_display = final_format if final_format is not None else 'None (no itemType/itemSubType)'
            logger.info(f"Row {row}: Searching for {len(row_isbns)} ISBNs with format '{format_display}': {', '.join(row_isbns)}")
            row_results = self.search_by_isbns(row_isbns, final_format)
            
            # Get the first match found (since all ISBNs are for the same work)
            oclc_number = None
            has_lcsh = False
            matched_isbn = None
            
            for isbn in row_isbns:
                if isbn in row_results:
                    result = row_results[isbn]
                    if isinstance(result, dict):
                        oclc_number = result.get('oclc_number')
                        has_lcsh = result.get('has_lcsh', False)
                    else:
                        # Backward compatibility for old format
                        oclc_number = result
                        has_lcsh = False
                    matched_isbn = isbn
                    break
            
            # Add OCLC number to the new column
            oclc_cell = worksheet.cell(row=row, column=oclc_col)
            oclc_cell.value = oclc_number
            
            # Add LCSH result to the new column
            lcsh_cell = worksheet.cell(row=row, column=lcsh_col)
            lcsh_cell.value = has_lcsh
            
            # Add other identifier to the new column (copy from input if available)
            other_id_cell = worksheet.cell(row=row, column=other_id_col)
            if other_id_col_input:
                other_id_value = worksheet.cell(row=row, column=other_id_col_input).value
                other_id_cell.value = other_id_value
            else:
                other_id_cell.value = ''
            
            # Update LCSH statistics if we found an OCLC number
            if oclc_number:
                if has_lcsh:
                    self.stats['lcsh_found'] += 1
                else:
                    self.stats['lcsh_not_found'] += 1
            
            # Update statistics
            self.stats['total_processed'] += 1
            if oclc_number:
                self.stats['successful_matches'] += 1
                logger.info(f"Row {row}: Found match with ISBN {matched_isbn} -> OCLC: {oclc_number}, LCSH: {has_lcsh}")
            else:
                self.stats['no_matches'] += 1
                logger.warning(f"Row {row}: No match found for ISBNs: {', '.join(row_isbns)}")
            
            # Add a small delay to be respectful to the API
            time.sleep(self.rate_limit_delay)
            
            # Progress update
            if (row - 1) % 50 == 0:
                elapsed_time = time.time() - start_time
                rate = (row - 1) / elapsed_time if elapsed_time > 0 else 0
                eta = (total_rows - row) / rate if rate > 0 else 0
                logger.info(f"Processed {row - 1}/{total_rows - 1} rows "
                          f"({rate:.1f} rows/sec, ETA: {eta/60:.1f} minutes)")
        
        # Save the updated Excel file
        logger.info(f"Saving results to: {output_file}")
        workbook.save(output_file)
        
        # Print final summary
        elapsed_time = time.time() - start_time
        self.print_summary(elapsed_time)

    def process_excel_file(self, input_file: str, output_file: str, create_backup: bool = True):
        """
        Process Excel file to add OCLC numbers using OR queries for ISBNs from the same row.

        Args:
            input_file: Path to input Excel file (.xlsx)
            output_file: Path to output Excel file
            create_backup: Whether to create a backup of the input file
        """
        try:
            if create_backup:
                self.create_backup(input_file)
            logger.info(f"Reading Excel file: {input_file}")
            workbook = openpyxl.load_workbook(input_file)
            self._process_workbook(workbook, output_file)
        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            raise

    def process_delimited_file(
        self,
        input_file: str,
        output_file: str,
        delimiter: str,
        create_backup: bool = True,
    ) -> None:
        """
        Process a CSV or TSV file (UTF-8 with optional BOM) and write results to an Excel file.

        Args:
            input_file: Path to input CSV or TSV
            output_file: Path to output .xlsx file
            delimiter: Field delimiter for csv.reader (comma or tab)
            create_backup: Whether to create a backup of the input file
        """
        try:
            if create_backup:
                self.create_backup(input_file)
            label = "TSV" if delimiter == "\t" else "CSV"
            logger.info(f"Reading {label} file: {input_file}")
            workbook = self._delimited_file_to_workbook(input_file, delimiter)
            self._process_workbook(workbook, output_file)
        except Exception as e:
            logger.error(f"Error processing delimited file: {e}")
            raise

    def extract_marc_data(self, marc_file: str) -> str:
        """
        Extract data from MARC file and create temporary Excel file.
        
        Args:
            marc_file: Path to MARC file
            
        Returns:
            Path to temporary Excel file with extracted data
        """
        import tempfile
        
        # Create temporary Excel file
        temp_dir = tempfile.gettempdir()
        temp_excel = os.path.join(temp_dir, f"marc_extracted_{int(time.time())}.xlsx")
        
        logger.info(f"Extracting data from MARC file: {marc_file}")
        
        # Use the marc_extractor module
        from marc_extractor import extract_marc_to_excel
        
        try:
            # Extract MARC data to temporary Excel file
            extract_marc_to_excel(marc_file, temp_excel, 'INFO')
            logger.info(f"Temporary Excel file created: {temp_excel}")
            return temp_excel
            
        except Exception as e:
            logger.error(f"Error extracting MARC data: {e}")
            raise

    def print_summary(self, elapsed_time: float):
        """Print processing summary."""
        logger.info("=" * 60)
        logger.info("PROCESSING COMPLETE!")
        logger.info("=" * 60)
        logger.info(f"Total records processed: {self.stats['total_processed']}")
        logger.info(f"Successful matches: {self.stats['successful_matches']}")
        logger.info(f"No matches found: {self.stats['no_matches']}")
        logger.info(f"Empty ISBNs: {self.stats['empty_isbns']}")
        logger.info(f"API errors: {self.stats['api_errors']}")
        logger.info(f"Success rate: {(self.stats['successful_matches']/self.stats['total_processed']*100):.1f}%")
        logger.info(f"LCSH subjects found: {self.stats['lcsh_found']}")
        logger.info(f"LCSH subjects not found: {self.stats['lcsh_not_found']}")
        if self.stats['successful_matches'] > 0:
            lcsh_rate = (self.stats['lcsh_found'] / self.stats['successful_matches'] * 100)
            logger.info(f"LCSH rate: {lcsh_rate:.1f}%")
        logger.info(f"Total time: {elapsed_time/60:.1f} minutes")
        logger.info(f"Average rate: {self.stats['total_processed']/elapsed_time:.1f} records/second")
        logger.info("=" * 60)


def detect_file_type(file_path: str) -> str:
    """
    Detect file type based on extension and content.
    
    Args:
        file_path: Path to the file
        
    Returns:
        File type: 'excel', 'csv', 'tsv', 'marc', or 'unknown'
    """
    path = Path(file_path)
    extension = path.suffix.lower()

    if extension in ['.xlsx', '.xls']:
        return 'excel'
    if extension == '.csv':
        return 'csv'
    if extension == '.tsv':
        return 'tsv'
    if extension in ['.mrc', '.marc']:
        return 'marc'
    return 'unknown'


def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(
        description="OCLC Record Matcher - Search OCLC Discovery Stratus API for ISBNs and add OCLC numbers to Excel, CSV, TSV, or MARC inputs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Use default Excel file
  python3 oclc_record_matcher.py
  
  # Process Excel file
  python3 oclc_record_matcher.py -i my_books.xlsx -o my_books_with_oclc.xlsx
  
  # Process MARC file
  python3 oclc_record_matcher.py -i catalog.mrc -o catalog_with_oclc.xlsx

  # Process CSV or TSV (UTF-8; output is always .xlsx)
  python3 oclc_record_matcher.py -i my_books.csv -o my_books_with_oclc.xlsx
  python3 oclc_record_matcher.py -i my_books.tsv -o my_books_with_oclc.xlsx
  
  # Process file without creating backup
  python3 oclc_record_matcher.py -i books.xlsx --no-backup
  
  # Use different log level
  python3 oclc_record_matcher.py -i books.xlsx --log-level DEBUG
        """
    )
    
    parser.add_argument(
        '-i', '--input',
        default='sampleData/recordsToMatch.xlsx',
        help='Input file path (.xlsx/.xls, .csv, .tsv, or MARC .mrc/.marc) (default: sampleData/recordsToMatch.xlsx)'
    )
    
    parser.add_argument(
        '-o', '--output',
        help='Output Excel file path (default: input_file_with_oclc.xlsx)'
    )
    
    parser.add_argument(
        '--no-backup',
        action='store_true',
        help='Skip creating backup of input file'
    )
    
    parser.add_argument(
        '--log-level',
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
        default='INFO',
        help='Set logging level (default: INFO)'
    )
    
    parser.add_argument(
        '--log-file',
        help='Custom log file path (default: oclc_matcher.log)'
    )
    
    parser.add_argument(
        '--no-api-logging',
        action='store_true',
        help='Disable detailed API request/response logging (reduces log verbosity)'
    )
    
    return parser.parse_args()


def setup_logging(log_level: str, log_file: str = None):
    """Setup logging configuration."""
    # Convert string to logging level
    numeric_level = getattr(logging, log_level.upper(), None)
    if not isinstance(numeric_level, int):
        raise ValueError(f'Invalid log level: {log_level}')
    
    # Configure logging
    handlers = [logging.StreamHandler()]
    if log_file:
        handlers.append(logging.FileHandler(log_file))
    else:
        handlers.append(logging.FileHandler('oclc_matcher.log'))
    
    logging.basicConfig(
        level=numeric_level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=handlers,
        force=True  # Override any existing configuration
    )


def main():
    """Main function to run the OCLC ISBN matcher."""
    
    # Parse command-line arguments
    args = parse_arguments()
    
    # Setup logging
    setup_logging(args.log_level, args.log_file)
    
    # Determine input and output files
    input_file = args.input
    if args.output:
        output_file = args.output
    else:
        # Generate output filename based on input filename
        base_name = os.path.splitext(input_file)[0]
        output_file = f"{base_name}_with_oclc.xlsx"
    
    # Check if input file exists
    if not os.path.exists(input_file):
        logger.error(f"Input file not found: {input_file}")
        logger.info("Please check the file path and try again.")
        sys.exit(1)
    
    # Detect file type
    file_type = detect_file_type(input_file)
    if file_type == 'unknown':
        logger.error(f"Unsupported file type: {input_file}")
        logger.info("Supported file types: .xlsx, .xls, .csv, .tsv, .mrc, .marc")
        sys.exit(1)
    
    # Check if output file already exists
    if os.path.exists(output_file):
        logger.warning(f"Output file already exists: {output_file}")
        response = input("Do you want to overwrite it? (y/N): ")
        if response.lower() not in ['y', 'yes']:
            logger.info("Operation cancelled by user.")
            sys.exit(0)
    
    # Display configuration
    logger.info("=" * 60)
    logger.info("OCLC ISBN Matcher Configuration")
    logger.info("=" * 60)
    logger.info(f"Input file: {input_file}")
    logger.info(f"Input file type: {file_type}")
    logger.info(f"Output file: {output_file}")
    logger.info(f"Create backup: {not args.no_backup}")
    logger.info(f"Log level: {args.log_level}")
    logger.info("=" * 60)
    
    # Create matcher instance with OAuth 2.0 authentication
    api_logging = not args.no_api_logging
    try:
        matcher = OCLCISBNMatcher(api_logging=api_logging)
    except ValueError as e:
        logger.error(f"Configuration error: {e}")
        logger.info("Please ensure OCLC_API_KEY and OCLC_API_SECRET are set in your .env file")
        logger.info("See .env.example for configuration details")
        sys.exit(1)
    
    # Process the file
    try:
        if file_type == 'marc':
            # Extract MARC data to temporary Excel file
            temp_excel = matcher.extract_marc_data(input_file)
            logger.info(f"Processing MARC file via temporary Excel: {temp_excel}")
            
            # Process the temporary Excel file
            matcher.process_excel_file(temp_excel, output_file, create_backup=False)
            
            # Clean up temporary file
            try:
                os.remove(temp_excel)
                logger.info("Temporary file cleaned up")
            except Exception as e:
                logger.warning(f"Could not remove temporary file {temp_excel}: {e}")

        elif file_type == 'csv':
            matcher.process_delimited_file(
                input_file, output_file, ",", create_backup=not args.no_backup
            )
        elif file_type == 'tsv':
            matcher.process_delimited_file(
                input_file, output_file, "\t", create_backup=not args.no_backup
            )
        else:  # excel
            matcher.process_excel_file(
                input_file, output_file, create_backup=not args.no_backup
            )
        
        # Print API statistics
        matcher.print_api_statistics()
        
        logger.info("Script completed successfully!")
    except Exception as e:
        logger.error(f"Script failed: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

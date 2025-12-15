#!/usr/bin/env python3
# Copyright 2024
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""
MARC Data Extractor

This script extracts specific fields from MARC21 bibliographic records and exports them to an Excel spreadsheet.

Extracted fields:
- 020$a: ISBN
- 245$a + 245$b: Title (combined, normalized)
- 100 or 110: Author (personal or corporate)
- 260$b or 264$b: Publisher
- 260$c or 264$c: Publication Date (normalized to 4-digit year)
- 300: Physical Description (all subfields)
- Format: Combined LDR 06 + 008 23 logic
"""

import argparse
import logging
import sys
from pathlib import Path
from typing import List, Dict, Optional

import pymarc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def setup_logging(log_level: str = 'INFO', log_file: Optional[str] = None) -> None:
    """
    Set up logging configuration.
    
    Args:
        log_level: Logging level (DEBUG, INFO, WARNING, ERROR)
        log_file: Optional log file path
    """
    level = getattr(logging, log_level.upper(), logging.INFO)
    
    handlers = [logging.StreamHandler(sys.stdout)]
    if log_file:
        handlers.append(logging.FileHandler(log_file))
    
    logging.basicConfig(
        level=level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=handlers
    )


def extract_isbn(record: pymarc.Record) -> str:
    """
    Extract ISBN from field 020 subfield a.
    
    Args:
        record: MARC record
        
    Returns:
        ISBN string or empty string if not found
    """
    # Get field 020 (ISBN)
    field_020 = record.get('020')
    if field_020 and field_020.get_subfields('a'):
        isbn = field_020.get_subfields('a')[0].strip()
        # Remove any additional text after the ISBN
        isbn = isbn.split()[0] if isbn else ''
        return isbn
    
    return ''


def extract_title(record: pymarc.Record) -> str:
    """
    Extract and normalize title from field 245 subfields a and b.
    Strips trailing punctuation and normalizes formatting.
    
    Args:
        record: MARC record
        
    Returns:
        Normalized combined title string or empty string if not found
    """
    field_245 = record.get('245')
    if not field_245:
        return ''
    
    title_parts = []
    
    # Get subfield a (title)
    if field_245.get_subfields('a'):
        title_a = field_245.get_subfields('a')[0].strip()
        # Remove trailing punctuation
        title_a = _normalize_title(title_a)
        title_parts.append(title_a)
    
    # Get subfield b (remainder of title)
    if field_245.get_subfields('b'):
        title_b = field_245.get_subfields('b')[0].strip()
        # Remove trailing punctuation
        title_b = _normalize_title(title_b)
        title_parts.append(title_b)
    
    return ' '.join(title_parts)


def _normalize_title(title_str: str) -> str:
    """
    Normalize title string by stripping trailing punctuation.
    
    Args:
        title_str: Raw title string
        
    Returns:
        Normalized title string with trailing punctuation removed
    """
    if not title_str:
        return ''
    
    # Remove trailing punctuation and extra whitespace
    # Common MARC title punctuation: /, :, ;, ., ,, =, etc.
    title_str = title_str.rstrip(' /:;.,=+').strip()
    
    return title_str


def extract_author(record: pymarc.Record) -> str:
    """
    Extract author from field 100 (personal name) or 110 (corporate name).
    
    Args:
        record: MARC record
        
    Returns:
        Author string or empty string if not found
    """
    # Try field 100 (personal name) first
    field_100 = record.get('100')
    if field_100:
        author_parts = []
        # Get subfield a (personal name)
        if field_100.get_subfields('a'):
            author_a = field_100.get_subfields('a')[0].strip()
            # Remove trailing punctuation
            author_a = author_a.rstrip(' ,')
            author_parts.append(author_a)
        
        # Get subfield d (dates)
        if field_100.get_subfields('d'):
            author_d = field_100.get_subfields('d')[0].strip()
            author_parts.append(f"({author_d})")
        
        return ' '.join(author_parts)
    
    # Try field 110 (corporate name) if no 100
    field_110 = record.get('110')
    if field_110:
        author_parts = []
        # Get subfield a (corporate name)
        if field_110.get_subfields('a'):
            author_a = field_110.get_subfields('a')[0].strip()
            # Remove trailing punctuation
            author_a = author_a.rstrip(' ,')
            author_parts.append(author_a)
        
        # Get subfield b (subordinate unit)
        if field_110.get_subfields('b'):
            author_b = field_110.get_subfields('b')[0].strip()
            author_parts.append(author_b)
        
        return ' '.join(author_parts)
    
    return ''


def extract_publisher(record: pymarc.Record) -> str:
    """
    Extract publisher from field 260$b or 264$b.
    
    Args:
        record: MARC record
        
    Returns:
        Publisher string or empty string if not found
    """
    # Try field 260 first (older format)
    field_260 = record.get('260')
    if field_260 and field_260.get_subfields('b'):
        publisher = field_260.get_subfields('b')[0].strip()
        # Remove trailing punctuation
        return publisher.rstrip(' ,')
    
    # Try field 264 (newer format)
    field_264 = record.get('264')
    if field_264 and field_264.get_subfields('b'):
        publisher = field_264.get_subfields('b')[0].strip()
        # Remove trailing punctuation
        return publisher.rstrip(' ,')
    
    return ''


def extract_publication_date(record: pymarc.Record) -> str:
    """
    Extract and normalize publication date from field 260$c or 264$c.
    Strips trailing punctuation and formats as 4-digit year.
    
    Args:
        record: MARC record
        
    Returns:
        Normalized publication date (4-digit year) or empty string if not found
    """
    import re
    
    # Try field 260 first (older format)
    field_260 = record.get('260')
    if field_260 and field_260.get_subfields('c'):
        date = field_260.get_subfields('c')[0].strip()
        return _normalize_publication_date(date)
    
    # Try field 264 (newer format)
    field_264 = record.get('264')
    if field_264 and field_264.get_subfields('c'):
        date = field_264.get_subfields('c')[0].strip()
        return _normalize_publication_date(date)
    
    return ''


def _normalize_publication_date(date_str: str) -> str:
    """
    Normalize publication date string to 4-digit year format.
    
    Args:
        date_str: Raw publication date string
        
    Returns:
        Normalized 4-digit year or empty string if no valid year found
    """
    import re
    
    if not date_str:
        return ''
    
    # Remove trailing punctuation and extra whitespace
    date_str = date_str.rstrip(' ,.;:').strip()
    
    # Look for 4-digit year patterns
    # Pattern 1: Direct 4-digit year (e.g., "2023", "2023.")
    year_match = re.search(r'\b(19|20)\d{2}\b', date_str)
    if year_match:
        return year_match.group()
    
    # Pattern 2: Year with month/day (e.g., "20231128", "2023-11-28")
    # Extract just the year part
    year_match = re.search(r'\b(19|20)\d{2}', date_str)
    if year_match:
        return year_match.group()
    
    # Pattern 3: Two-digit year (e.g., "23", "99")
    # Convert to 4-digit assuming 20xx for 00-99, 19xx for 00-99
    two_digit_match = re.search(r'\b(\d{2})\b', date_str)
    if two_digit_match:
        year = int(two_digit_match.group())
        if year <= 99:
            # Assume 20xx for years 00-99
            return f"20{year:02d}"
    
    # If no year pattern found, return empty string
    return ''


def extract_form(record: pymarc.Record) -> str:
    """
    Extract form of material from 008 field position 23.
    
    Args:
        record: MARC record
        
    Returns:
        Form character or empty string if not found
    """
    field_008 = record.get('008')
    if field_008 and len(field_008.data) >= 24:  # Position 23 is index 23 (0-based)
        return field_008.data[23]
    return ''


def extract_type(record: pymarc.Record) -> str:
    """
    Extract type of record from Leader position 06.
    
    Args:
        record: MARC record
        
    Returns:
        Type character or empty string if not found
    """
    if record.leader:
        leader_str = str(record.leader)
        if len(leader_str) >= 7:  # Position 06 is index 6 (0-based)
            return leader_str[6]
    return ''


def extract_physical_description(record: pymarc.Record) -> str:
    """
    Extract physical description from field 300.
    
    Args:
        record: MARC record
        
    Returns:
        Physical description string or empty string if not found
    """
    field_300 = record.get('300')
    if field_300:
        # Use the field's value() method which combines all subfields
        description = field_300.value()
        if description:
            # Clean up extra spaces
            description = ' '.join(description.split())
            return description
    return ''


def extract_other_identifier(record: pymarc.Record) -> str:
    """
    Extract other identifier from 024 subfield a.
    
    Args:
        record: MARC record
        
    Returns:
        Other identifier string or empty string if not found
    """
    field_024 = record.get('024')
    if field_024:
        subfield_a = field_024.get_subfields('a')
        if subfield_a:
            return subfield_a[0]  # Return first subfield a value
    return ''


def extract_format(record: pymarc.Record, physical_description: str = '') -> str:
    """
    Extract format based on LDR position 06 and 008 position 23.
    Check physical description last for specific overrides.
    
    Logic:
    - If LDR 06 is g then set format to video
    - If LDR 06 is i then set format to audiobook
    - If LDR 06 is j then set format to music
    - If LDR 06 is a and 008 23 is d set format to book-largeprint
    - If LDR 06 is a and 008 23 is s set format to book-digital
    - If LDR is a and 008 23 is any other value set format to book-print
    - If physical description contains 'computer' and 'game', override to 'game'
    - If physical description contains 'computer' (but not 'game'), override to 'compfile'
    
    Args:
        record: MARC record
        physical_description: Physical description string to check for computer content
        
    Returns:
        Format string or empty string if not found
    """
    # Check physical description first for specific overrides
    if physical_description:
        desc_lower = physical_description.lower()
        # Check for computer + game combination first (more specific)
        if 'computer' in desc_lower and 'game' in desc_lower:
            return 'game'
        elif 'computer' in desc_lower:
            return 'compfile'
    
    # Get LDR position 06
    ldr_06 = extract_type(record)
    if not ldr_06:
        return ''
    
    # Get 008 position 23
    form_23 = extract_form(record)
    
    # Apply format logic
    if ldr_06 == 'g':
        return 'video'
    elif ldr_06 == 'i':
        return 'audiobook'
    elif ldr_06 == 'j':
        return 'music'
    elif ldr_06 == 'a':
        if form_23 == 'd':
            return 'book-largeprint'
        elif form_23 == 's':
            return 'book-digital'
        else:
            return 'book-print'
    
    return ''


def process_marc_file(input_file: str) -> List[Dict[str, str]]:
    """
    Process MARC file and extract required fields.
    
    Args:
        input_file: Path to MARC file
        
    Returns:
        List of dictionaries containing extracted data
    """
    records_data = []
    
    try:
        with open(input_file, 'rb') as file:
            reader = pymarc.MARCReader(file)
            
            for i, record in enumerate(reader, 1):
                try:
                    # Extract all required fields
                    isbn = extract_isbn(record)
                    title = extract_title(record)
                    author = extract_author(record)
                    publisher = extract_publisher(record)
                    publication_date = extract_publication_date(record)
                    physical_description = extract_physical_description(record)
                    other_identifier = extract_other_identifier(record)
                    format_str = extract_format(record, physical_description)
                    
                    record_data = {
                        'ISBN': isbn,
                        'Title': title,
                        'Author': author,
                        'Publisher': publisher,
                        'Publication Date': publication_date,
                        'Physical Description': physical_description,
                        'Other Identifier': other_identifier,
                        'Format': format_str
                    }
                    
                    records_data.append(record_data)
                    
                    if i % 100 == 0:
                        logging.info(f"Processed {i} records...")
                        
                except Exception as e:
                    logging.error(f"Error processing record {i}: {e}")
                    continue
                    
    except Exception as e:
        logging.error(f"Error reading MARC file: {e}")
        raise
    
    logging.info(f"Successfully processed {len(records_data)} records")
    return records_data


def create_excel_file(records_data: List[Dict[str, str]], output_file: str) -> None:
    """
    Create Excel file with extracted MARC data.
    
    Args:
        records_data: List of dictionaries containing extracted data
        output_file: Path to output Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "MARC Data"
    
    # Define headers
    headers = ['ISBN', 'Title', 'Author', 'Publisher', 'Publication Date', 'Physical Description', 'Format']
    
    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write data
    for row, record in enumerate(records_data, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=record.get(header, ''))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    logging.info(f"Excel file saved: {output_file}")


def extract_marc_to_excel(input_file: str, output_file: str, log_level: str = 'INFO') -> None:
    """
    Extract MARC data to Excel file (module function).
    
    Args:
        input_file: Path to input MARC file
        output_file: Path to output Excel file
        log_level: Logging level
    """
    # Set up logging
    setup_logging(log_level)
    
    # Validate input file
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file does not exist: {input_file}")
    
    if not input_path.suffix.lower() in ['.mrc', '.marc']:
        logging.warning(f"Input file doesn't have .mrc or .marc extension: {input_file}")
    
    # Validate output file
    output_path = Path(output_file)
    if not output_path.suffix.lower() == '.xlsx':
        logging.warning(f"Output file doesn't have .xlsx extension: {output_file}")
    
    logging.info("Starting MARC data extraction...")
    logging.info(f"Input file: {input_file}")
    logging.info(f"Output file: {output_file}")
    
    # Process MARC file
    records_data = process_marc_file(input_file)
    
    if not records_data:
        logging.warning("No records found in MARC file")
        return
    
    # Create Excel file
    create_excel_file(records_data, output_file)
    
    logging.info("MARC extraction completed successfully!")
    
    # Print summary
    print(f"\n{'='*60}")
    print("EXTRACTION SUMMARY")
    print(f"{'='*60}")
    print(f"Total records processed: {len(records_data)}")
    print(f"Output file: {output_file}")
    print(f"{'='*60}")


def main():
    """Main function to run the MARC extractor."""
    parser = argparse.ArgumentParser(
        description='Extract MARC data to Excel spreadsheet',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python marc_extractor.py -i sample.mrc -o output.xlsx
  python marc_extractor.py -i sample.mrc -o output.xlsx --log-level DEBUG
  python marc_extractor.py -i sample.mrc -o output.xlsx --log-file marc_extraction.log
        """
    )
    
    parser.add_argument(
        '-i', '--input',
        required=True,
        help='Input MARC file (.mrc)'
    )
    
    parser.add_argument(
        '-o', '--output',
        required=True,
        help='Output Excel file (.xlsx)'
    )
    
    parser.add_argument(
        '--log-level',
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
        default='INFO',
        help='Set logging level (default: INFO)'
    )
    
    parser.add_argument(
        '--log-file',
        help='Log file path (optional)'
    )
    
    args = parser.parse_args()
    
    # Set up logging
    setup_logging(args.log_level, args.log_file)
    
    # Validate input file
    input_path = Path(args.input)
    if not input_path.exists():
        logging.error(f"Input file does not exist: {args.input}")
        sys.exit(1)
    
    if not input_path.suffix.lower() in ['.mrc', '.marc']:
        logging.warning(f"Input file doesn't have .mrc or .marc extension: {args.input}")
    
    # Validate output file
    output_path = Path(args.output)
    if not output_path.suffix.lower() == '.xlsx':
        logging.warning(f"Output file doesn't have .xlsx extension: {args.output}")
    
    try:
        # Use the module function
        extract_marc_to_excel(args.input, args.output, args.log_level)
        
    except Exception as e:
        logging.error(f"Error during extraction: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()

#!/usr/bin/env python3
"""
MARC Field Analyzer

This script analyzes a MARC21 bibliographic file to determine the most common fields.

Author: AI Assistant
License: Apache License 2.0
"""

import argparse
import logging
import sys
from collections import Counter, defaultdict
from pathlib import Path

import pymarc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def setup_logging(log_level: str, log_file: str = None) -> None:
    """
    Set up logging configuration.
    
    Args:
        log_level: Logging level (DEBUG, INFO, WARNING, ERROR)
        log_file: Optional log file path
    """
    level = getattr(logging, log_level.upper())
    
    handlers = [logging.StreamHandler(sys.stdout)]
    if log_file:
        handlers.append(logging.FileHandler(log_file))
    
    logging.basicConfig(
        level=level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=handlers
    )


def analyze_marc_fields(input_file: str) -> dict:
    """
    Analyze MARC file to count field occurrences.
    
    Args:
        input_file: Path to input MARC file
        
    Returns:
        Dictionary containing field statistics
    """
    field_counts = Counter()
    subfield_counts = defaultdict(Counter)
    control_field_counts = Counter()
    leader_positions = defaultdict(Counter)
    
    total_records = 0
    
    try:
        with open(input_file, 'rb') as file:
            reader = pymarc.MARCReader(file)
            
            for i, record in enumerate(reader, 1):
                if i % 100 == 0:
                    logging.info(f"Processed {i} records...")
                
                total_records += 1
                
                # Count data fields
                for field in record.get_fields():
                    if field.tag.isdigit() and int(field.tag) <= 9:
                        # Control field (001-009)
                        control_field_counts[field.tag] += 1
                    else:
                        # Data field (010 and above, or non-numeric)
                        field_counts[field.tag] += 1
                        
                        # Count subfields
                        if hasattr(field, 'subfields'):
                            for subfield in field.subfields:
                                if len(subfield) == 2:  # subfield code + value
                                    subfield_code = subfield[0]
                                    subfield_counts[field.tag][subfield_code] += 1
                
                # Analyze leader positions
                if record.leader:
                    leader_str = str(record.leader)
                    for pos, char in enumerate(leader_str):
                        leader_positions[pos][char] += 1
                        
    except Exception as e:
        logging.error(f"Error analyzing MARC file: {e}")
        raise
    
    logging.info(f"Successfully analyzed {total_records} records")
    
    return {
        'total_records': total_records,
        'field_counts': field_counts,
        'subfield_counts': subfield_counts,
        'control_field_counts': control_field_counts,
        'leader_positions': leader_positions
    }


def create_analysis_excel(analysis_data: dict, output_file: str) -> None:
    """
    Create Excel file with field analysis results.
    
    Args:
        analysis_data: Dictionary containing analysis results
        output_file: Path to output Excel file
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    ws_fields = wb.create_sheet("Field Counts")
    ws_control = wb.create_sheet("Control Fields")
    ws_subfields = wb.create_sheet("Subfield Counts")
    ws_leader = wb.create_sheet("Leader Analysis")
    ws_summary = wb.create_sheet("Summary")
    
    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Summary sheet
    ws_summary.cell(row=1, column=1, value="MARC Field Analysis Summary")
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    ws_summary.cell(row=3, column=1, value="Total Records:")
    ws_summary.cell(row=3, column=2, value=analysis_data['total_records'])
    
    ws_summary.cell(row=4, column=1, value="Total Data Fields:")
    ws_summary.cell(row=4, column=2, value=sum(analysis_data['field_counts'].values()))
    
    ws_summary.cell(row=5, column=1, value="Total Control Fields:")
    ws_summary.cell(row=5, column=2, value=sum(analysis_data['control_field_counts'].values()))
    
    ws_summary.cell(row=6, column=1, value="Unique Data Field Tags:")
    ws_summary.cell(row=6, column=2, value=len(analysis_data['field_counts']))
    
    ws_summary.cell(row=7, column=1, value="Unique Control Field Tags:")
    ws_summary.cell(row=7, column=2, value=len(analysis_data['control_field_counts']))
    
    # Data fields sheet
    ws_fields.cell(row=1, column=1, value="Field Tag").font = header_font
    ws_fields.cell(row=1, column=1).fill = header_fill
    ws_fields.cell(row=1, column=2, value="Count").font = header_font
    ws_fields.cell(row=1, column=2).fill = header_fill
    ws_fields.cell(row=1, column=3, value="Percentage").font = header_font
    ws_fields.cell(row=1, column=3).fill = header_fill
    
    total_data_fields = sum(analysis_data['field_counts'].values())
    for i, (field_tag, count) in enumerate(analysis_data['field_counts'].most_common(), 2):
        ws_fields.cell(row=i, column=1, value=field_tag)
        ws_fields.cell(row=i, column=2, value=count)
        ws_fields.cell(row=i, column=3, value=f"{count/total_data_fields*100:.2f}%")
    
    # Control fields sheet
    ws_control.cell(row=1, column=1, value="Control Field Tag").font = header_font
    ws_control.cell(row=1, column=1).fill = header_fill
    ws_control.cell(row=1, column=2, value="Count").font = header_font
    ws_control.cell(row=1, column=2).fill = header_fill
    ws_control.cell(row=1, column=3, value="Percentage").font = header_font
    ws_control.cell(row=1, column=3).fill = header_fill
    
    total_control_fields = sum(analysis_data['control_field_counts'].values())
    for i, (field_tag, count) in enumerate(analysis_data['control_field_counts'].most_common(), 2):
        ws_control.cell(row=i, column=1, value=field_tag)
        ws_control.cell(row=i, column=2, value=count)
        ws_control.cell(row=i, column=3, value=f"{count/total_control_fields*100:.2f}%")
    
    # Subfields sheet
    ws_subfields.cell(row=1, column=1, value="Field Tag").font = header_font
    ws_subfields.cell(row=1, column=1).fill = header_fill
    ws_subfields.cell(row=1, column=2, value="Subfield Code").font = header_font
    ws_subfields.cell(row=1, column=2).fill = header_fill
    ws_subfields.cell(row=1, column=3, value="Count").font = header_font
    ws_subfields.cell(row=1, column=3).fill = header_fill
    
    row = 2
    for field_tag, subfield_counter in analysis_data['subfield_counts'].items():
        for subfield_code, count in subfield_counter.most_common():
            ws_subfields.cell(row=row, column=1, value=field_tag)
            ws_subfields.cell(row=row, column=2, value=subfield_code)
            ws_subfields.cell(row=row, column=3, value=count)
            row += 1
    
    # Leader analysis sheet
    ws_leader.cell(row=1, column=1, value="Position").font = header_font
    ws_leader.cell(row=1, column=1).fill = header_fill
    ws_leader.cell(row=1, column=2, value="Character").font = header_font
    ws_leader.cell(row=1, column=2).fill = header_fill
    ws_leader.cell(row=1, column=3, value="Count").font = header_font
    ws_leader.cell(row=1, column=3).fill = header_fill
    ws_leader.cell(row=1, column=4, value="Percentage").font = header_font
    ws_leader.cell(row=1, column=4).fill = header_fill
    
    row = 2
    for pos, char_counter in analysis_data['leader_positions'].items():
        total_chars = sum(char_counter.values())
        for char, count in char_counter.most_common():
            ws_leader.cell(row=row, column=1, value=pos)
            ws_leader.cell(row=row, column=2, value=char)
            ws_leader.cell(row=row, column=3, value=count)
            ws_leader.cell(row=row, column=4, value=f"{count/total_chars*100:.2f}%")
            row += 1
    
    # Auto-adjust column widths
    for ws in [ws_fields, ws_control, ws_subfields, ws_leader, ws_summary]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    logging.info(f"Analysis Excel file saved: {output_file}")


def print_summary(analysis_data: dict) -> None:
    """
    Print analysis summary to console.
    
    Args:
        analysis_data: Dictionary containing analysis results
    """
    print("\n" + "="*60)
    print("MARC FIELD ANALYSIS SUMMARY")
    print("="*60)
    print(f"Total records analyzed: {analysis_data['total_records']}")
    print(f"Total data fields: {sum(analysis_data['field_counts'].values())}")
    print(f"Total control fields: {sum(analysis_data['control_field_counts'].values())}")
    print(f"Unique data field tags: {len(analysis_data['field_counts'])}")
    print(f"Unique control field tags: {len(analysis_data['control_field_counts'])}")
    
    print("\n" + "-"*40)
    print("TOP 20 MOST COMMON DATA FIELDS")
    print("-"*40)
    for i, (field_tag, count) in enumerate(analysis_data['field_counts'].most_common(20), 1):
        percentage = count / sum(analysis_data['field_counts'].values()) * 100
        print(f"{i:2d}. {field_tag}: {count:4d} occurrences ({percentage:5.2f}%)")
    
    print("\n" + "-"*40)
    print("ALL CONTROL FIELDS")
    print("-"*40)
    for field_tag, count in analysis_data['control_field_counts'].most_common():
        percentage = count / sum(analysis_data['control_field_counts'].values()) * 100
        print(f"{field_tag}: {count:4d} occurrences ({percentage:5.2f}%)")
    
    print("\n" + "-"*40)
    print("LEADER POSITION ANALYSIS")
    print("-"*40)
    for pos in sorted(analysis_data['leader_positions'].keys()):
        char_counter = analysis_data['leader_positions'][pos]
        most_common = char_counter.most_common(1)[0]
        print(f"Position {pos:2d}: '{most_common[0]}' ({most_common[1]} occurrences)")


def main():
    """Main function."""
    parser = argparse.ArgumentParser(
        description="Analyze MARC file to determine most common fields",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python marc_field_analyzer.py -i sample.mrc -o analysis.xlsx
  python marc_field_analyzer.py -i sample.mrc -o analysis.xlsx --log-level DEBUG
  python marc_field_analyzer.py -i sample.mrc -o analysis.xlsx --log-file analysis.log
        """
    )
    
    parser.add_argument(
        '-i', '--input',
        required=True,
        help='Input MARC file (.mrc)'
    )
    
    parser.add_argument(
        '-o', '--output',
        required=True,
        help='Output Excel file (.xlsx)'
    )
    
    parser.add_argument(
        '--log-level',
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
        default='INFO',
        help='Set logging level (default: INFO)'
    )
    
    parser.add_argument(
        '--log-file',
        help='Log file path (optional)'
    )
    
    args = parser.parse_args()
    
    # Set up logging
    setup_logging(args.log_level, args.log_file)
    
    # Validate input file
    input_path = Path(args.input)
    if not input_path.exists():
        logging.error(f"Input file does not exist: {args.input}")
        sys.exit(1)
    
    if not input_path.suffix.lower() in ['.mrc', '.marc']:
        logging.warning(f"Input file doesn't have .mrc or .marc extension: {args.input}")
    
    # Validate output file
    output_path = Path(args.output)
    if not output_path.suffix.lower() == '.xlsx':
        logging.warning(f"Output file doesn't have .xlsx extension: {args.output}")
    
    try:
        logging.info("Starting MARC field analysis...")
        logging.info(f"Input file: {args.input}")
        logging.info(f"Output file: {args.output}")
        
        # Analyze MARC file
        analysis_data = analyze_marc_fields(args.input)
        
        # Create Excel output
        create_analysis_excel(analysis_data, args.output)
        
        # Print summary
        print_summary(analysis_data)
        
        logging.info("MARC field analysis completed successfully!")
        
    except Exception as e:
        logging.error(f"Error during analysis: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
